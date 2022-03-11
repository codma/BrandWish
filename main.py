from typing import List, Any

from openpyxl import Workbook, load_workbook

# write_wb = Workbook()
# write_wb.save("/Users/jinjoolee/Desktop/Project/python/BrandWish.xlsx")
wb = load_workbook("/Users/jinjoolee/Desktop/Project/python/BrandWish.xlsx")
# print(wb.sheetnames)
write_ws = wb['Sheet']

wish_wb = load_workbook("/Users/jinjoolee/Desktop/Project/python/wishlist.xlsx")
wish_ws = wish_wb['Sheet']

wish_brand_wb = load_workbook("/Users/jinjoolee/Desktop/Project/python/wishlist_brand.xlsx")
wish_brand_ws = wish_brand_wb['Sheet']

brand_wb = load_workbook("/Users/jinjoolee/Desktop/Project/python/brand.xlsx")
brand_ws = brand_wb['Sheet']

all_values1 = []
all_values2 = []
all_values3 = []
result_values = []

cellCount = 0
for row in wish_ws.rows:
    row_value1 = []
    for cell in row:
        if cellCount == 0:
            row_value1.append(cell.value)
        if cellCount == 1:
            row_value1.append(cell.value)
        if cellCount == 2:
            row_value1.append(cell.value)
        cellCount += 1
    cellCount = 0
    all_values1.append(row_value1)
# wish_id = 0	user_id = 1	type = 2


for row2 in wish_brand_ws.rows:
    row_value2 = []
    for cell2 in row2:
        row_value2.append(cell2.value)
    all_values2.append(row_value2)
# id	wish_id	brand_id	is_checked	is_deleted	created_date	modified_date	deleted_date

for row3 in brand_ws.rows:
    row_value3 = []
    for cell3 in row3:
        if cellCount == 0:
            row_value3.append(cell3.value)
        if cellCount == 1:
            row_value3.append(cell3.value)
        if cellCount == 2:
            row_value3.append(cell3.value)
        cellCount += 1
    cellCount = 0
    all_values3.append(row_value3)
#brand_id = 0	name_eng = 1	name_kor = 2	designer_type

cellCount2 = 0

rowCount = 0
rowCount2 = 0

for row4 in all_values1:
    results = []
    cellCount = 0
    if rowCount == 0:
        rowCount += 1
        continue
    for cell4 in row4:
        rowCount2 = 0
        if cellCount == 0:
            for row5 in all_values2:
                if rowCount2 == 0:
                    rowCount2 += 1
                    continue
                cellCount2 = 0
                for cell5 in row5:
                    if cellCount2 == 1:
                        if cell4 == cell5:
                            results.append(str(row4[0]))
                            results.append(str(row4[1]))
                            results.append(str(row5[2]))
                            results.append(str(row5[3]))
                            results.append(str(row5[4]))
                            results.append(str(row5[5]))
                            results.append(row5[6])
                            results.append(row5[7])
                    cellCount2 += 1
                if len(results) != 0:
                    write_ws.append(results)
                    # 테스트
                    # wb.save("/Users/jinjoolee/Desktop/Project/python/BrandWish/xlsx")
                    results = []
    write_ws.append(results)

wb.save("/Users/jinjoolee/Desktop/Project/python/BrandWish.xlsx")